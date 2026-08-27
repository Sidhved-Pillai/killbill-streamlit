import io
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


SUMMARY_COLUMNS = [
    "Sr. No.",
    "Invoice Date",
    "Invoice Number",
    "Truck Number",
    "From",
    "Customer Name",
    "Location",
    "Vehicle",
    "Case",
    "Jar",
    "Freight",
    "Toll",
    "Total Freight",
]

SOURCE_DATE_FORMATS = (
    "%d-%b-%y",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%Y-%m-%d",
)


def _parse_invoice_date(value):
    if value is None or pd.isna(value):
        return pd.NaT
    for date_format in SOURCE_DATE_FORMATS:
        try:
            return pd.Timestamp(datetime.strptime(str(value).strip(), date_format))
        except ValueError:
            continue
    return pd.to_datetime(value, errors="coerce", dayfirst=True)


def _group_key(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def _numeric_series(dataframe, column):
    if column not in dataframe.columns:
        return pd.Series(0, index=dataframe.index, dtype=float)
    return pd.to_numeric(dataframe[column], errors="coerce").fillna(0)


def build_billing_statement(dataframe):
    """Organize every reviewed invoice into monthly billing-statement rows."""
    if dataframe is None or dataframe.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS + ["_bill_month"])

    working = dataframe.copy()
    working["_parsed_date"] = working.get("Date", "").map(_parse_invoice_date)
    working["_bill_month"] = working["_parsed_date"].dt.to_period("M")
    working["_truck_key"] = working.get("Vehicle No.", "").map(_group_key)
    working["_from_key"] = working.get("From", "").map(_group_key)
    working["_to_key"] = working.get("To", "").map(_group_key)

    group_columns = ["_bill_month", "_truck_key", "_from_key", "_to_key"]
    organized = working.sort_values(
        group_columns + ["_parsed_date", "Invoice No."],
        kind="stable",
        na_position="last",
    )
    freight = _numeric_series(organized, "Freight Charge")
    summary = pd.DataFrame(
        {
            "Invoice Date": organized["_parsed_date"].dt.strftime("%d-%b-%y"),
            "Invoice Number": organized.get("Invoice No.", ""),
            "Truck Number": organized.get("Vehicle No.", ""),
            "From": organized.get("From", ""),
            "Customer Name": organized.get("Customer Name", ""),
            "Location": organized.get("To", ""),
            "Vehicle": "9MT",
            "Case": _numeric_series(organized, "Case"),
            "Jar": _numeric_series(organized, "Jar"),
            "Freight": freight,
            "Toll": 0,
            "Total Freight": freight,
            "_bill_month": organized["_bill_month"],
        }
    ).reset_index(drop=True)
    summary.insert(
        0,
        "Sr. No.",
        summary.groupby("_bill_month", dropna=False).cumcount() + 1,
    )
    return summary


def bill_number_for_month(month):
    """Build the monthly bill number shown in the supplied summary template."""
    timestamp = month.to_timestamp() if isinstance(month, pd.Period) else pd.Timestamp(month)
    financial_year_start = timestamp.year if timestamp.month >= 4 else timestamp.year - 1
    return (
        f"BIPL-L{timestamp.month:02d}1/"
        f"{financial_year_start % 100:02d}-{(financial_year_start + 1) % 100:02d}"
    )


def _safe_sheet_title(month):
    if pd.isna(month):
        return "Undated"
    return month.to_timestamp().strftime("%b %Y")


def _write_month_sheet(worksheet, month_data, month):
    title = "UBIQUITY TRANSTECH PVT LTD."
    if pd.isna(month):
        subtitle = "Trips Summary for invoices without a valid date"
    else:
        month_label = month.to_timestamp().strftime("%B %Y")
        subtitle = (
            f"Trips Summary for the month of {month_label}, "
            f"Bill No.:-{bill_number_for_month(month)}"
        )

    worksheet.merge_cells("A1:M1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(name="Times New Roman", size=20, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells("A2:M2")
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(bold=True)
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="000000")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column_number, header in enumerate(SUMMARY_COLUMNS, start=1):
        cell = worksheet.cell(row=3, column=column_number, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = table_border

    first_data_row = 4
    for offset, (_, row) in enumerate(month_data.iterrows()):
        row_number = first_data_row + offset
        for column_number, header in enumerate(SUMMARY_COLUMNS, start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=row[header])
            cell.border = table_border
            cell.alignment = Alignment(
                horizontal="center" if header in {"Sr. No.", "Vehicle", "Case", "Jar"} else "left",
                vertical="center",
            )

    minimum_entry_rows = 20
    total_row = first_data_row + max(len(month_data), minimum_entry_rows)
    for row_number in range(first_data_row + len(month_data), total_row):
        for column_number in range(1, len(SUMMARY_COLUMNS) + 1):
            worksheet.cell(row=row_number, column=column_number).border = table_border

    worksheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
    total_label = worksheet.cell(row=total_row, column=1, value="Grand Total =======>")
    total_label.font = Font(bold=True)
    total_label.alignment = Alignment(horizontal="center")
    for column_number in range(1, len(SUMMARY_COLUMNS) + 1):
        worksheet.cell(row=total_row, column=column_number).border = table_border

    for column_number, header in ((11, "Freight"), (12, "Toll"), (13, "Total Freight")):
        total = float(pd.to_numeric(month_data[header], errors="coerce").fillna(0).sum())
        total_cell = worksheet.cell(row=total_row, column=column_number, value=total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0"

    signature_row = total_row + 3
    worksheet.merge_cells(
        start_row=signature_row,
        start_column=9,
        end_row=signature_row,
        end_column=13,
    )
    worksheet.cell(
        row=signature_row,
        column=9,
        value="For : UBIQUITY TRANSTECH PRIVATE LIMITED",
    ).font = Font(bold=True)
    worksheet.merge_cells(
        start_row=signature_row + 4,
        start_column=9,
        end_row=signature_row + 4,
        end_column=13,
    )
    worksheet.cell(
        row=signature_row + 4,
        column=9,
        value="Authorised Signatory",
    ).font = Font(bold=True)

    widths = [7, 15, 20, 17, 14, 30, 24, 10, 9, 9, 15, 12, 18]
    for column_number, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = width
    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = False
    worksheet.print_area = f"A1:M{signature_row + 5}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def build_billing_statement_workbook(summary):
    """Create one template-formatted worksheet for each invoice month."""
    if summary is None or summary.empty:
        raise ValueError("Billing statement has no invoices to export.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    months = list(summary["_bill_month"].drop_duplicates())
    for month in months:
        worksheet = workbook.create_sheet(_safe_sheet_title(month))
        if pd.isna(month):
            month_data = summary.loc[summary["_bill_month"].isna(), SUMMARY_COLUMNS]
        else:
            month_data = summary.loc[summary["_bill_month"] == month, SUMMARY_COLUMNS]
        _write_month_sheet(worksheet, month_data, month)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def billing_statement_filename(summary):
    months = list(summary["_bill_month"].drop_duplicates())
    if len(months) == 1 and not pd.isna(months[0]):
        suffix = months[0].to_timestamp().strftime("%B_%Y")
    elif len(months) == 1:
        suffix = "Undated"
    else:
        suffix = "Multiple_Months"
    return f"Billing_Statement_{suffix}.xlsx"
