import io
import unittest
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from excel_export import (
    EXPORT_COLUMNS,
    FREIGHT_NUMBER_FORMAT,
    build_excel_workbook,
    excel_export_filename,
)


class ExcelExportTests(unittest.TestCase):
    def setUp(self):
        self.dataframe = pd.DataFrame(
            [
                {
                    "Date": "24-Jul-26",
                    "Invoice No.": "INV-001",
                    "Vehicle No.": "MH-04 AB 1234",
                    "From": "Thane",
                    "Customer Code": "MUMC001",
                    "Customer Name": "Sample Customer",
                    "To": "Mumbai",
                    "Vehicle Type": "9MT",
                    "Case": 12,
                    "Jar": 3,
                    "Freight Charge": 4508.64,
                    "Lookup Status": "✅ Matched",
                }
            ]
        )
        self.workbook = load_workbook(io.BytesIO(build_excel_workbook(self.dataframe)))
        self.worksheet = self.workbook.active

    def test_workbook_contains_exact_columns_and_values(self):
        expected_headers = [header for _, header in EXPORT_COLUMNS]
        actual_headers = [cell.value for cell in self.worksheet[1]]
        self.assertEqual(actual_headers, expected_headers)

        expected_values = [
            self.dataframe.iloc[0][source_column]
            for source_column, _ in EXPORT_COLUMNS
        ]
        actual_values = [cell.value for cell in self.worksheet[2]]
        self.assertEqual(actual_values, expected_values)

    def test_workbook_formatting(self):
        self.assertEqual(self.worksheet.freeze_panes, "A2")
        self.assertTrue(all(cell.font.bold for cell in self.worksheet[1]))
        self.assertTrue(
            all(cell.fill.fill_type == "solid" for cell in self.worksheet[1])
        )

        header_positions = {
            cell.value: cell.column for cell in self.worksheet[1]
        }
        for header in ("Vehicle Type", "Case", "Jar"):
            self.assertEqual(
                self.worksheet.cell(2, header_positions[header]).alignment.horizontal,
                "center",
            )

        self.assertEqual(
            self.worksheet.cell(2, header_positions["Freight Charge"]).number_format,
            FREIGHT_NUMBER_FORMAT,
        )
        self.assertEqual(
            self.worksheet.cell(2, header_positions["Lookup Status"]).number_format,
            "@",
        )
        self.assertTrue(
            all(
                self.worksheet.column_dimensions[cell.column_letter].width > 0
                for cell in self.worksheet[1]
            )
        )

    def test_filename_uses_required_timestamp(self):
        filename = excel_export_filename(datetime(2026, 7, 24, 3, 30))
        self.assertEqual(filename, "Bisleri_Billing_2026-07-24_03-30.xlsx")


if __name__ == "__main__":
    unittest.main()
