import io
import unittest

import pandas as pd
from openpyxl import load_workbook

from billing_statement import (
    SUMMARY_COLUMNS,
    billing_statement_filename,
    bill_number_for_month,
    build_billing_statement,
    build_billing_statement_workbook,
)


class BillingStatementTests(unittest.TestCase):
    def setUp(self):
        self.dataframe = pd.DataFrame(
            [
                {
                    "Date": "01-May-26",
                    "Invoice No.": "INV-001",
                    "Vehicle No.": "MH04HD3996",
                    "From": "Vasai",
                    "Customer Name": "First Customer",
                    "To": "Bhiwandi",
                    "Vehicle Type": "6MT",
                    "Case": 50,
                    "Jar": 120,
                    "Freight Charge": 4509,
                },
                {
                    "Date": "08-May-26",
                    "Invoice No.": "INV-002",
                    "Vehicle No.": "MH04HD3996",
                    "From": "Vasai",
                    "Customer Name": "Second Customer",
                    "To": "Bhiwandi",
                    "Vehicle Type": "12MT",
                    "Case": 25,
                    "Jar": 40,
                    "Freight Charge": 4509,
                },
                {
                    "Date": "09-May-26",
                    "Invoice No.": "INV-003",
                    "Vehicle No.": "MH04HD3996",
                    "From": "Thane",
                    "Customer Name": "Different Route",
                    "To": "Bhiwandi",
                    "Vehicle Type": "9MT",
                    "Case": 10,
                    "Jar": 0,
                    "Freight Charge": 3844,
                },
                {
                    "Date": "10-May-26",
                    "Invoice No.": "INV-004",
                    "Vehicle No.": "OTHER-TRUCK",
                    "From": "Vasai",
                    "Customer Name": "Different Truck",
                    "To": "Bhiwandi",
                    "Vehicle Type": "9MT",
                    "Case": 5,
                    "Jar": 0,
                    "Freight Charge": 3844,
                },
            ]
        )

    def test_includes_every_invoice_and_organizes_without_aggregating(self):
        summary = build_billing_statement(self.dataframe)

        self.assertEqual(
            list(summary["Invoice Number"]),
            ["INV-003", "INV-001", "INV-002", "INV-004"],
        )
        self.assertEqual(len(summary), len(self.dataframe))
        self.assertEqual(list(summary["Case"]), [10, 50, 25, 5])
        self.assertEqual(list(summary["Jar"]), [0, 120, 40, 0])
        self.assertEqual(list(summary["Vehicle"]), ["9MT"] * 4)
        self.assertEqual(list(summary["Total Freight"]), [3844, 4509, 4509, 3844])

    def test_same_route_in_different_months_keeps_every_invoice(self):
        other_month = self.dataframe.iloc[[0]].copy()
        other_month["Date"] = "01-Jun-26"
        data = pd.concat([self.dataframe.iloc[[0]], other_month], ignore_index=True)

        summary = build_billing_statement(data)

        self.assertEqual(len(summary), 2)
        self.assertEqual(summary["Sr. No."].tolist(), [1, 1])

    def test_workbook_matches_template_title_headers_totals_and_signature(self):
        summary = build_billing_statement(self.dataframe)
        workbook = load_workbook(
            io.BytesIO(build_billing_statement_workbook(summary))
        )
        worksheet = workbook["May 2026"]

        self.assertEqual(worksheet["A1"].value, "UBIQUITY TRANSTECH PVT LTD.")
        self.assertEqual(
            worksheet["A2"].value,
            "Trips Summary for the month of May 2026, Bill No.:-BIPL-L051/26-27",
        )
        self.assertEqual(
            [worksheet.cell(3, column).value for column in range(1, 14)],
            SUMMARY_COLUMNS,
        )
        self.assertEqual(worksheet["C4"].value, "INV-003")
        self.assertEqual(worksheet["H4"].value, "9MT")
        self.assertEqual(worksheet["I4"].value, 10)
        self.assertEqual(worksheet["J4"].value, 0)
        self.assertEqual(worksheet["K24"].value, 16706)
        self.assertEqual(worksheet["M24"].value, 16706)
        self.assertEqual(
            worksheet["I27"].value,
            "For : UBIQUITY TRANSTECH PRIVATE LIMITED",
        )

    def test_bill_number_and_filename_follow_invoice_month(self):
        month = pd.Period("2026-05", freq="M")
        summary = build_billing_statement(self.dataframe)

        self.assertEqual(bill_number_for_month(month), "BIPL-L051/26-27")
        self.assertEqual(
            billing_statement_filename(summary),
            "Billing_Statement_May_2026.xlsx",
        )

    def test_undated_review_row_is_still_included_and_exportable(self):
        undated = self.dataframe.iloc[[0]].copy()
        undated["Date"] = ""

        summary = build_billing_statement(undated)
        workbook = load_workbook(
            io.BytesIO(build_billing_statement_workbook(summary))
        )

        self.assertEqual(len(summary), 1)
        self.assertEqual(summary.iloc[0]["Invoice Number"], "INV-001")
        self.assertIn("Undated", workbook.sheetnames)
        self.assertEqual(
            billing_statement_filename(summary),
            "Billing_Statement_Undated.xlsx",
        )

    def test_multi_month_workbook_opens_largest_sheet_and_keeps_every_row(self):
        june = self.dataframe.iloc[[0]].copy()
        june["Date"] = "30-Jun-26"
        july = self.dataframe.iloc[[1, 2, 3]].copy()
        july["Date"] = ["01-Jul-26", "02-Jul-26", "03-Jul-26"]
        source = pd.concat([june, july], ignore_index=True)

        summary = build_billing_statement(source)
        workbook = load_workbook(
            io.BytesIO(build_billing_statement_workbook(summary))
        )

        self.assertEqual(workbook.sheetnames, ["Jun 2026", "Jul 2026"])
        self.assertEqual(workbook.active.title, "Jul 2026")
        exported_invoice_numbers = []
        for worksheet in workbook.worksheets:
            exported_invoice_numbers.extend(
                worksheet.cell(row_number, 3).value
                for row_number in range(4, worksheet.max_row + 1)
                if worksheet.cell(row_number, 3).value
            )
        self.assertCountEqual(
            exported_invoice_numbers,
            source["Invoice No."].tolist(),
        )


if __name__ == "__main__":
    unittest.main()
