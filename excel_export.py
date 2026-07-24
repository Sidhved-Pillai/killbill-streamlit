import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    ("Date", "Date"),
    ("Invoice No.", "Invoice No"),
    ("Vehicle No.", "Vehicle No"),
    ("From", "From"),
    ("Customer Code", "Customer Code"),
    ("Customer Name", "Customer Name"),
    ("To", "To"),
    ("Vehicle Type", "Vehicle Type"),
    ("Case", "Case"),
    ("Jar", "Jar"),
    ("Freight Charge", "Freight Charge"),
    ("Lookup Status", "Lookup Status"),
]

HEADER_FILL = "E7E6E6"
FREIGHT_NUMBER_FORMAT = "[$₹-en-IN] #,##,##0.00"


def _excel_value(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value.item() if hasattr(value, "item") else value


def build_excel_workbook(dataframe):
    """Return an .xlsx file containing the table's current displayed data."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Bisleri Billing"

    for column_number, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)

    for row_number, (_, row) in enumerate(dataframe.iterrows(), start=2):
        for column_number, (source_column, _) in enumerate(EXPORT_COLUMNS, start=1):
            value = _excel_value(row[source_column]) if source_column in dataframe.columns else None
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            if isinstance(value, str):
                # Prevent text that begins with "=" from being converted into a formula.
                cell.data_type = "s"

    worksheet.freeze_panes = "A2"

    centered_columns = {"Vehicle Type", "Case", "Jar"}
    for column_number, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        if header in centered_columns:
            for cell in worksheet.iter_cols(
                min_col=column_number,
                max_col=column_number,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.alignment = Alignment(horizontal="center")

        if header == "Freight Charge":
            for cell in worksheet.iter_cols(
                min_col=column_number,
                max_col=column_number,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = FREIGHT_NUMBER_FORMAT

        if header == "Lookup Status":
            for cell in worksheet.iter_cols(
                min_col=column_number,
                max_col=column_number,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "@"

    for column_number in range(1, worksheet.max_column + 1):
        longest_value = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in worksheet[get_column_letter(column_number)]
        )
        worksheet.column_dimensions[get_column_letter(column_number)].width = longest_value + 2

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_export_filename(now=None):
    timestamp = now or datetime.now()
    return f"Bisleri_Billing_{timestamp:%Y-%m-%d_%H-%M}.xlsx"
